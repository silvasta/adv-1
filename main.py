from math import dist
from openpyxl import Workbook

def main():
        # Aufgabe 1
    file = open("Koordinatenliste.dat")
        # a)
    # ansatz1(file)
    # ansatz2(file)
        # b)
    # print_koordinaten(file)
        # c)
    #distanz1(file)
        # d)
    #create_excel(file)
        # e)
    distanz2(file)

    file.close()


def ansatz1(file):
    for count, line in enumerate(file):
        pass
    print("Insgesamt hat es {} Datensätze".format(count+1))

def ansatz2(file):
    count = (len(file.readlines()))
    print("Insgesamt hat es {} Datensätze".format(count))

def print_koordinaten(file):
    for line in file:
        print(line)

def distanz1(file):
    sum = 0.0
    point1 = file.readline().split()
    point1 = (int(point1[0]),int(point1[1]))
    for line in file.readlines():
        point2 = line.split()
        point2 = (int(point2[0]),int(point2[1]))
        sum += dist(point1,point2)
        point1 = point2
    sum_meter = int(round(sum, 0))
    print("Die Summe aller Abstände in Metern ist {}".format(sum_meter))

def create_excel(file):
    # create workbook and activate sheet
    wb = Workbook()
    ws = wb.active

    # create header
    ws.cell(row=1, column=1).value = "x1"
    ws.cell(row=1, column=2).value = "x2"
    ws.cell(row=1, column=3).value = "dist x1"
    ws.cell(row=1, column=4).value = "dist x2"
    ws.cell(row=1, column=5).value = "dist"


    row = 2
    point1 = file.readline().split()
    ws.cell(row=row, column=1).value = point1[0]
    ws.cell(row=row, column=2).value = point1[1]
    
    for line in file.readlines():
        row += 1
        point2 = line.split()
        ws.cell(row=row, column=1).value = point2[0]
        ws.cell(row=row, column=2).value = point2[1]
        ws.cell(row=row, column=3).value = "=A{}-A{}".format(row-1, row)
        ws.cell(row=row, column=4).value = "=B{}-B{}".format(row-1, row)
        ws.cell(row=row, column=5).value = "=SQRT(C{}^2+D{}^2)".format(row,row)


    ws.cell(row=row+1, column=5).value = "=SUM(E2:E{})".format(row)

    wb.save(filename="check_in_excel.xlsx")

def distanz2(file):
    sum = 0.0
    file = file.read().split("\n")
    point_before = False
    for point in file:
        point = point.split() # trenne die Koordinaten
        if len(point) != 0: # verhindere leere Zeilen        
            point = (int(point[0]),int(point[1])) # konvertiere Liste mit Strings zu Tuple        
            if not point_before: # erste Distanz gibt es nicht
                point_before = point
            else:
                sum += dist(point,point_before) 
                point_before = point

    sum_meter = int(round(sum, 0))
    print("Die Summe aller Abstände in Metern ist {}".format(sum_meter))



if __name__ == '__main__':
  main()