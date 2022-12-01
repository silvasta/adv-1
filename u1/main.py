from math import dist
from openpyxl import Workbook
import csv

def main():
        # Aufgabe 1
    file = open("Koordinatenliste.dat")
        # a)
    # ansatz1(file)
    # ansatz2(file)
        # b)
    # print_koordinaten(file)
        # c)
    #distanz1(file)
        # d)
    #create_excel(file)
        # e)
    #distanz2(file)

    file.close()

        # Aufgabe 2
        # a)
    # read_csv()
        # b)
    # read_dat()
        # c)
    #change_csv()
        # d)
    # excel öffnet sich mit Spalten vertauscht
        # e)
    sort_csv()
    


def ansatz1(file):
    for count, line in enumerate(file):
        pass
    print("Insgesamt hat es {} Datensätze".format(count+1))

def ansatz2(file):
    count = (len(file.readlines()))
    print("Insgesamt hat es {} Datensätze".format(count))

def print_koordinaten(file):
    for line in file:
        print(line)

def distanz1(file):
    sum = 0.0
    point1 = file.readline().split()
    point1 = (int(point1[0]),int(point1[1]))
    for line in file.readlines():
        point2 = line.split()
        point2 = (int(point2[0]),int(point2[1]))
        sum += dist(point1,point2)
        point1 = point2
    sum_meter = int(round(sum, 0))
    print("Die Summe aller Abstände in Metern ist {}".format(sum_meter))

def create_excel(file):
    # create workbook and activate sheet
    wb = Workbook()
    ws = wb.active

    # create header
    ws.cell(row=1, column=1).value = "x1"
    ws.cell(row=1, column=2).value = "x2"
    ws.cell(row=1, column=3).value = "dist x1"
    ws.cell(row=1, column=4).value = "dist x2"
    ws.cell(row=1, column=5).value = "dist"


    row = 2
    point1 = file.readline().split()
    ws.cell(row=row, column=1).value = point1[0]
    ws.cell(row=row, column=2).value = point1[1]
    
    for line in file.readlines():
        row += 1
        point2 = line.split()
        ws.cell(row=row, column=1).value = point2[0]
        ws.cell(row=row, column=2).value = point2[1]
        ws.cell(row=row, column=3).value = "=A{}-A{}".format(row-1, row)
        ws.cell(row=row, column=4).value = "=B{}-B{}".format(row-1, row)
        ws.cell(row=row, column=5).value = "=SQRT(C{}^2+D{}^2)".format(row,row)


    ws.cell(row=row+1, column=5).value = "=SUM(E2:E{})".format(row)

    wb.save(filename="check_in_excel.xlsx")

def distanz2(file):
    sum = 0.0
    file = file.read().split("\n")
    point_before = False
    for point in file:
        point = point.split() # trenne die Koordinaten
        if len(point) != 0: # verhindere leere Zeilen        
            point = (int(point[0]),int(point[1])) # konvertiere Liste mit Strings zu Tuple        
            if not point_before: # erste Distanz gibt es nicht
                point_before = point
            else:
                sum += dist(point,point_before) 
                point_before = point

    sum_meter = int(round(sum, 0))
    print("Die Summe aller Abstände in Metern ist {}".format(sum_meter))

def read_csv():
    file = open("CH_PLZ_GDE.csv")
    csvreader = csv.reader(file, delimiter = ";")
    for line in csvreader:
        print(line[1])

def read_dat():
    file = open("CH_PLZ_GDE.dat")
    csvreader = csv.reader(file, delimiter = " ")
    for line in csvreader:
        print(line[1])

def change_csv():
    input_file = open("CH_PLZ_GDE.csv")
    csvreader = csv.reader(input_file, delimiter = ";")
    output_file = open("CH_GDE_PLZ.csv", "wt")
    csvwriter = csv.writer(output_file, delimiter=";", lineterminator="\n")

    for line in csvreader:
        csvwriter.writerow([line[1], line[0]])

def sort_csv():
    input_file = open("CH_PLZ_GDE.csv")
    csvreader = csv.reader(input_file, delimiter = ";")
    output_file = open("sorted_CH_GDE_PLZ.csv", "wt")
    csvwriter = csv.writer(output_file, delimiter=";", lineterminator="\n")
    gde_list = []

        # sort with dict
    # for line in csvreader:
    #     gde_list.append({"PLZ" : line[0], "Gemeinde" : line[1]})
    # gde_sorted = sorted(gde_list, key= lambda d:d["Gemeinde"])
    # for line in gde_sorted:
    #     csvwriter.writerow([line["Gemeinde"], line["PLZ"]])

        # sort with list
    for line in csvreader:
        gde_list.append([line[0], line[1]])
    gde_sorted = sorted(gde_list, key = lambda d:d[1])
    for line in gde_sorted:
        csvwriter.writerow([line[1], line[0]])


if __name__ == '__main__':
  main()